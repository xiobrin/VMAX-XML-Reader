# -*- coding: cp1252 -*-

import subprocess

filepath="C:\\users\\kjensb\\sym0158.bat"
#filepath="C:\\users\\kjensb\\sym0219.bat"
#filepath="C:\\users\\kjensb\\sym0227.bat"
#filepath="C:\\users\\kjensb\\sym0390.bat"

print 'Running SYMCLI commands....'
p = subprocess.Popen(filepath, shell=True, stdout = subprocess.PIPE)

stdout, stderr = p.communicate()
if p.returncode == 0:
    print 'Success...'

    import xml.etree.ElementTree as ET
    f = open('c:\\users\\kjensb\\file.xml', "r")
    e = open('c:\\users\\kjensb\\engines.xml', "r")
    s = open('c:\\users\\kjensb\\srp.xml', "r")


    if e != None:
        tree = ET.parse(e)
        root = tree.getroot()

    # counting total disk in the array
    engine_counter = 0
    found_engine = 0
    l=[]

    # iterating thru xml
    for child in root.iter():
        #print child
        #print 'tag', child.tag
        #print 'text', child.text
        disk_group = child.find('Disk_Group_Engine_info')
        engine_num = child.find('engine_num')
        spare_num = child.find('total_spare_disks')
        if engine_num != None:
            found_engine = 1
            #print engine_num.text
            if engine_num.text != 'N/A':
                engine_counter = engine_counter + 1
        if spare_num != None:
            if found_engine == 1:
                if spare_num.text != 'N/A':
                    spare_counter = int(spare_num.text)
                    while spare_counter > 0:
                        l.append('1')
                        spare_counter = spare_counter - 1
                    #print 'Engine ' + str(engine_counter) + ' has ' + str(spare_num.text) + ' spare disks.'
                    #print l
                    found_engine = 0
                

    if s != None:
        tree = ET.parse(s)
        root = tree.getroot()

    allocated_cap = 0
    subscribed_cap = 0
    usable_cap = 0

    # iterating thru xml
    for child in root.iter():
        #print child
        #print 'tag', child.tag
        #print 'text', child.text

        subscribed = child.find('total_subscribed_capacity_terabytes')
        if subscribed != None:
            subscribed_cap = float(subscribed.text)
        allocated = child.find('total_allocated_capacity_terabytes')
        if allocated != None:
            allocated_cap = float(allocated.text)
        usable = child.find('total_usable_capacity_terabytes')
        if usable != None:
            usable_cap = float(usable.text)
            

    from openpyxl import load_workbook
    wb2 = load_workbook('C:\\Users\\kjensb\\calculator\\CALCULATOR TEMPLATE v2.0.xlsx')

    #from openpyxl.styles import PatternFill
    #fill = PatternFill(fill_type=None,start_color='FFFFFFFF',end_color='FF000000')
    #failed_fill = GradientFill(stop=("FFBB00", "FFBB00"))

    #assign the active sheet
    ws = wb2.active

    subscription = 0
    subscription = round(float(subscribed_cap/usable_cap),2)
    #print subscribed_cap
    #print usable_cap
    #print subscription
    
    ws['C7'] = engine_counter
    ws['C10'] = subscription
    ws['C14'] = allocated_cap
    ws['C15'] = subscribed_cap

    # Python types will automatically be converted
    #import datetime
    #ws['A2'] = datetime.datetime.now()

    # set env variables for now
    DAECount = 16
    DAECounter = 1
    EngineCount = 6
    DAEnumber = 16
    lastDAE = "DF-1C"

    # variable used to balance drives over all DAE
    BalanceTracker = 0

    # assuming DAE has 120 slots (0-119)
    remainingslots = 119

    # template list for Engine1 in the spreadsheet
    ENG1list = ['F4','G4','H4','I4','J4','K4','L4','M4','N4','O4','P4','Q4','R4','S4','T4','U4','V4','W4','X4','Y4','Z4','AA4','AB4','AC4',]

    # template list for Engine2 in the spreadsheet
    ENG2list = ['F7','G7','H7','I7','J7','K7','L7','M7','N7','O7','P7','Q7','R7','S7','T7','U7','V7','W7','X7','Y7','Z7','AA7','AB7','AC7',]

    # template list for Engine3 in the spreadsheet
    ENG3list = ['AF4','AG4','AH4','AI4','AJ4','AK4','AL4','AM4','AN4','AO4','AP4','AQ4','AR4','AS4','AT4','AU4','AV4','AW4','AX4','AY4','AZ4','BA4','BB4','BC4',]

    # template list for Engine4 in the spreadsheet
    ENG4list = ['AF7','AG7','AH7','AI7','AJ7','AK7','AL7','AM7','AN7','AO7','AP7','AQ7','AR7','AS7','AT7','AU7','AV7','AW7','AX7','AY7','AZ7','BA7','BB7','BC7',]

    # template list for Engine5 in the spreadsheet
    ENG5list = ['BF4','BG4','BH4','BI4','BJ4','BK4','BL4','BM4','BN4','BO4','BP4','BQ4','BR4','BS4','BT4','BU4','BV4','BW4','BX4','BY4','BZ4','CA4','CB4','CC4',]

    # template list for Engine6 in the spreadsheet
    ENG6list = ['BF7','BG7','BH7','BI7','BJ7','BK7','BL7','BM7','BN7','BO7','BP7','BQ7','BR7','BS7','BT7','BU7','BV7','BW7','BX7','BY7','BZ7','CA7','CB7','CC7',]

    # template list for Engine7 in the spreadsheet
    ENG7list = ['CF4','CG4','CH4','CI4','CJ4','CK4','CL4','CM4','CN4','CO4','CP4','CQ4','CR4','CS4','CT4','CU4','CV4','CW4','CX4','CY4','CZ4','DA4','DB4','DC4',]

    # template list for Engine8 in the spreadsheet
    ENG8list = ['CF7','CG7','CH7','CI7','CJ7','CK7','CL7','CM7','CN7','CO7','CP7','CQ7','CR7','CS7','CT7','CU7','CV7','CW7','CX7','CY7','CZ7','DA7','DB7','DC7',]


    # template list for DAE1 in the spreadsheet
    DAE1list = ['F9','G9','H9','I9','J9','K9','L9','M9','N9','O9','P9','Q9','R9','S9','T9','U9','V9','W9','X9','Y9','Z9','AA9','AB9','AC9',
                'F10','G10','H10','I10','J10','K10','L10','M10','N10','O10','P10','Q10','R10','S10','T10','U10','V10','W10','X10','Y10','Z10','AA10','AB10','AC10',
                'F11','G11','H11','I11','J11','K11','L11','M11','N11','O11','P11','Q11','R11','S11','T11','U11','V11','W11','X11','Y11','Z11','AA11','AB11','AC11',
                'F12','G12','H12','I12','J12','K12','L12','M12','N12','O12','P12','Q12','R12','S12','T12','U12','V12','W12','X12','Y12','Z12','AA12','AB12','AC12',
                'F13','G13','H13','I13','J13','K13','L13','M13','N13','O13','P13','Q13','R13','S13','T13','U13','V13','W13','X13','Y13','Z13','AA13','AB13','AC13',]
    # template list for DAE2 in the spreadsheet
    DAE2list = ['F15','G15','H15','I15','J15','K15','L15','M15','N15','O15','P15','Q15','R15','S15','T15','U15','V15','W15','X15','Y15','Z15','AA15','AB15','AC15',
                'F16','G16','H16','I16','J16','K16','L16','M16','N16','O16','P16','Q16','R16','S16','T16','U16','V16','W16','X16','Y16','Z16','AA16','AB16','AC16',
                'F17','G17','H17','I17','J17','K17','L17','M17','N17','O17','P17','Q17','R17','S17','T17','U17','V17','W17','X17','Y17','Z17','AA17','AB17','AC17',
                'F18','G18','H18','I18','J18','K18','L18','M18','N18','O18','P18','Q18','R18','S18','T18','U18','V18','W18','X18','Y18','Z18','AA18','AB18','AC18',
                'F19','G19','H19','I19','J19','K19','L19','M19','N19','O19','P19','Q19','R19','S19','T19','U19','V19','W19','X19','Y19','Z19','AA19','AB19','AC19',]
    # template list for DAE3 in the spreadsheet
    DAE3list = ['F21','G21','H21','I21','J21','K21','L21','M21','N21','O21','P21','Q21','R21','S21','T21','U21','V21','W21','X21','Y21','Z21','AA21','AB21','AC21',
                'F22','G22','H22','I22','J22','K22','L22','M22','N22','O22','P22','Q22','R22','S22','T22','U22','V22','W22','X22','Y22','Z22','AA22','AB22','AC22',
                'F23','G23','H23','I23','J23','K23','L23','M23','N23','O23','P23','Q23','R23','S23','T23','U23','V23','W23','X23','Y23','Z23','AA23','AB23','AC23',
                'F24','G24','H24','I24','J24','K24','L24','M24','N24','O24','P24','Q24','R24','S24','T24','U24','V24','W24','X24','Y24','Z24','AA24','AB24','AC24',
                'F25','G25','H25','I25','J25','K25','L25','M25','N25','O25','P25','Q25','R25','S25','T25','U25','V25','W25','X25','Y25','Z25','AA25','AB25','AC25',]
    # template list for DAE4 in the spreadsheet
    DAE4list = ['F27','G27','H27','I27','J27','K27','L27','M27','N27','O27','P27','Q27','R27','S27','T27','U27','V27','W27','X27','Y27','Z27','AA27','AB27','AC27',
                'F28','G28','H28','I28','J28','K28','L28','M28','N28','O28','P28','Q28','R28','S28','T28','U28','V28','W28','X28','Y28','Z28','AA28','AB28','AC28',
                'F29','G29','H29','I29','J29','K29','L29','M29','N29','O29','P29','Q29','R29','S29','T29','U29','V29','W29','X29','Y29','Z29','AA29','AB29','AC29',
                'F30','G30','H30','I30','J30','K30','L30','M30','N30','O30','P30','Q30','R30','S30','T30','U30','V30','W30','X30','Y30','Z30','AA30','AB30','AC30',
                'F31','G31','H31','I31','J31','K31','L31','M31','N31','O31','P31','Q31','R31','S31','T31','U31','V31','W31','X31','Y31','Z31','AA31','AB31','AC31',]

    # template list for DAE5 in the spreadsheet
    DAE5list = ['AF9','AG9','AH9','AI9','AJ9','AK9','AL9','AM9','AN9','AO9','AP9','AQ9','AR9','AS9','AT9','AU9','AV9','AW9','AX9','AY9','AZ9','BA9','BB9','BC9',
                'AF10','AG10','AH10','AI10','AJ10','AK10','AL10','AM10','AN10','AO10','AP10','AQ10','AR10','AS10','AT10','AU10','AV10','AW10','AX10','AY10','AZ10','BA10','BB10','BC10',
                'AF11','AG11','AH11','AI11','AJ11','AK11','AL11','AM11','AN11','AO11','AP11','AQ11','AR11','AS11','AT11','AU11','AV11','AW11','AX11','AY11','AZ11','BA11','BB11','BC11',
                'AF12','AG12','AH12','AI12','AJ12','AK12','AL12','AM12','AN12','AO12','AP12','AQ12','AR12','AS12','AT12','AU12','AV12','AW12','AX12','AY12','AZ12','BA12','BB12','BC12',
                'AF13','AG13','AH13','AI13','AJ13','AK13','AL13','AM13','AN13','AO13','AP13','AQ13','AR13','AS13','AT13','AU13','AV13','AW13','AX13','AY13','AZ13','BA13','BB13','BC13',]

    # template list for DAE6 in the spreadsheet
    DAE6list = ['AF15','AG15','AH15','AI15','AJ15','AK15','AL15','AM15','AN15','AO15','AP15','AQ15','AR15','AS15','AT15','AU15','AV15','AW15','AX15','AY15','AZ15','BA15','BB15','BC15',
                'AF16','AG16','AH16','AI16','AJ16','AK16','AL16','AM16','AN16','AO16','AP16','AQ16','AR16','AS16','AT16','AU16','AV16','AW16','AX16','AY16','AZ16','BA16','BB16','BC16',
                'AF17','AG17','AH17','AI17','AJ17','AK17','AL17','AM17','AN17','AO17','AP17','AQ17','AR17','AS17','AT17','AU17','AV17','AW17','AX17','AY17','AZ17','BA17','BB17','BC17',
                'AF18','AG18','AH18','AI18','AJ18','AK18','AL18','AM18','AN18','AO18','AP18','AQ18','AR18','AS18','AT18','AU18','AV18','AW18','AX18','AY18','AZ18','BA18','BB18','BC18',
                'AF19','AG19','AH19','AI19','AJ19','AK19','AL19','AM19','AN19','AO19','AP19','AQ19','AR19','AS19','AT19','AU19','AV19','AW19','AX19','AY19','AZ19','BA19','BB19','BC19',]

    # template list for DAE7 in the spreadsheet
    DAE7list = ['AF21','AG21','AH21','AI21','AJ21','AK21','AL21','AM21','AN21','AO21','AP21','AQ21','AR21','AS21','AT21','AU21','AV21','AW21','AX21','AY21','AZ21','BA21','BB21','BC21',
                'AF22','AG22','AH22','AI22','AJ22','AK22','AL22','AM22','AN22','AO22','AP22','AQ22','AR22','AS22','AT22','AU22','AV22','AW22','AX22','AY22','AZ22','BA22','BB22','BC22',
                'AF23','AG23','AH23','AI23','AJ23','AK23','AL23','AM23','AN23','AO23','AP23','AQ23','AR23','AS23','AT23','AU23','AV23','AW23','AX23','AY23','AZ23','BA23','BB23','BC23',
                'AF24','AG24','AH24','AI24','AJ24','AK24','AL24','AM24','AN24','AO24','AP24','AQ24','AR24','AS24','AT24','AU24','AV24','AW24','AX24','AY24','AZ24','BA24','BB24','BC24',
                'AF25','AG25','AH25','AI25','AJ25','AK25','AL25','AM25','AN25','AO25','AP25','AQ25','AR25','AS25','AT25','AU25','AV25','AW25','AX25','AY25','AZ25','BA25','BB25','BC25',]

    # template list for DAE8 in the spreadsheet
    DAE8list = ['AF27','AG27','AH27','AI27','AJ27','AK27','AL27','AM27','AN27','AO27','AP27','AQ27','AR27','AS27','AT27','AU27','AV27','AW27','AX27','AY27','AZ27','BA27','BB27','BC27',
                'AF28','AG28','AH28','AI28','AJ28','AK28','AL28','AM28','AN28','AO28','AP28','AQ28','AR28','AS28','AT28','AU28','AV28','AW28','AX28','AY28','AZ28','BA28','BB28','BC28',
                'AF29','AG29','AH29','AI29','AJ29','AK29','AL29','AM29','AN29','AO29','AP29','AQ29','AR29','AS29','AT29','AU29','AV29','AW29','AX29','AY29','AZ29','BA29','BB29','BC29',
                'AF30','AG30','AH30','AI30','AJ30','AK30','AL30','AM30','AN30','AO30','AP30','AQ30','AR30','AS30','AT30','AU30','AV30','AW30','AX30','AY30','AZ30','BA30','BB30','BC30',
                'AF31','AG31','AH31','AI31','AJ31','AK31','AL31','AM31','AN31','AO31','AP31','AQ31','AR31','AS31','AT31','AU31','AV31','AW31','AX31','AY31','AZ31','BA31','BB31','BC31',]

    # template list for DAE9 in the spreadsheet
    DAE9list = ['BF9','BG9','BH9','BI9','BJ9','BK9','BL9','BM9','BN9','BO9','BP9','BQ9','BR9','BS9','BT9','BU9','BV9','BW9','BX9','BY9','BZ9','CA9','CB9','CC9',
                'BF10','BG10','BH10','BI10','BJ10','BK10','BL10','BM10','BN10','BO10','BP10','BQ10','BR10','BS10','BT10','BU10','BV10','BW10','BX10','BY10','BZ10','CA10','CB10','CC10',
                'BF11','BG11','BH11','BI11','BJ11','BK11','BL11','BM11','BN11','BO11','BP11','BQ11','BR11','BS11','BT11','BU11','BV11','BW11','BX11','BY11','BZ11','CA11','CB11','CC11',
                'BF12','BG12','BH12','BI12','BJ12','BK12','BL12','BM12','BN12','BO12','BP12','BQ12','BR12','BS12','BT12','BU12','BV12','BW12','BX12','BY12','BZ12','CA12','CB12','CC12',
                'BF13','BG13','BH13','BI13','BJ13','BK13','BL13','BM13','BN13','BO13','BP13','BQ13','BR13','BS13','BT13','BU13','BV13','BW13','BX13','BY13','BZ13','CA13','CB13','CC13',]

    # template list for DAE10 in the spreadsheet
    DAE10list = ['BF15','BG15','BH15','BI15','BJ15','BK15','BL15','BM15','BN15','BO15','BP15','BQ15','BR15','BS15','BT15','BU15','BV15','BW15','BX15','BY15','BZ15','CA15','CB15','CC15',
                'BF16','BG16','BH16','BI16','BJ16','BK16','BL16','BM16','BN16','BO16','BP16','BQ16','BR16','BS16','BT16','BU16','BV16','BW16','BX16','BY16','BZ16','CA16','CB16','CC16',
                'BF17','BG17','BH17','BI17','BJ17','BK17','BL17','BM17','BN17','BO17','BP17','BQ17','BR17','BS17','BT17','BU17','BV17','BW17','BX17','BY17','BZ17','CA17','CB17','CC17',
                'BF18','BG18','BH18','BI18','BJ18','BK18','BL18','BM18','BN18','BO18','BP18','BQ18','BR18','BS18','BT18','BU18','BV18','BW18','BX18','BY18','BZ18','CA18','CB18','CC18',
                'BF19','BG19','BH19','BI19','BJ19','BK19','BL19','BM19','BN19','BO19','BP19','BQ19','BR19','BS19','BT19','BU19','BV19','BW19','BX19','BY19','BZ19','CA19','CB19','CC19',]

    # template list for DAE11 in the spreadsheet
    DAE11list = ['BF21','BG21','BH21','BI21','BJ21','BK21','BL21','BM21','BN21','BO21','BP21','BQ21','BR21','BS21','BT21','BU21','BV21','BW21','BX21','BY21','BZ21','CA21','CB21','CC21',
                'BF22','BG22','BH22','BI22','BJ22','BK22','BL22','BM22','BN22','BO22','BP22','BQ22','BR22','BS22','BT22','BU22','BV22','BW22','BX22','BY22','BZ22','CA22','CB22','CC22',
                'BF23','BG23','BH23','BI23','BJ23','BK23','BL23','BM23','BN23','BO23','BP23','BQ23','BR23','BS23','BT23','BU23','BV23','BW23','BX23','BY23','BZ23','CA23','CB23','CC23',
                'BF24','BG24','BH24','BI24','BJ24','BK24','BL24','BM24','BN24','BO24','BP24','BQ24','BR24','BS24','BT24','BU24','BV24','BW24','BX24','BY24','BZ24','CA24','CB24','CC24',
                'BF25','BG25','BH25','BI25','BJ25','BK25','BL25','BM25','BN25','BO25','BP25','BQ25','BR25','BS25','BT25','BU25','BV25','BW25','BX25','BY25','BZ25','CA25','CB25','CC25',]

    # template list for DAE12 in the spreadsheet
    DAE12list = ['BF27','BG27','BH27','BI27','BJ27','BK27','BL27','BM27','BN27','BO27','BP27','BQ27','BR27','BS27','BT27','BU27','BV27','BW27','BX27','BY27','BZ27','CA27','CB27','CC27',
                'BF28','BG28','BH28','BI28','BJ28','BK28','BL28','BM28','BN28','BO28','BP28','BQ28','BR28','BS28','BT28','BU28','BV28','BW28','BX28','BY28','BZ28','CA28','CB28','CC28',
                'BF29','BG29','BH29','BI29','BJ29','BK29','BL29','BM29','BN29','BO29','BP29','BQ29','BR29','BS29','BT29','BU29','BV29','BW29','BX29','BY29','BZ29','CA29','CB29','CC29',
                'BF30','BG30','BH30','BI30','BJ30','BK30','BL30','BM30','BN30','BO30','BP30','BQ30','BR30','BS30','BT30','BU30','BV30','BW30','BX30','BY30','BZ30','CA30','CB30','CC30',
                'BF31','BG31','BH31','BI31','BJ31','BK31','BL31','BM31','BN31','BO31','BP31','BQ31','BR31','BS31','BT31','BU31','BV31','BW31','BX31','BY31','BZ31','CA31','CB31','CC31',]

    # template list for DAE13 in the spreadsheet
    DAE13list = ['CF9','CG9','CH9','CI9','CJ9','CK9','CL9','CM9','CN9','CO9','CP9','CQ9','CR9','CS9','CT9','CU9','CV9','CW9','CX9','CY9','CZ9','DA9','DB9','DC9',
                'CF10','CG10','CH10','CI10','CJ10','CK10','CL10','CM10','CN10','CO10','CP10','CQ10','CR10','CS10','CT10','CU10','CV10','CW10','CX10','CY10','CZ10','DA10','DB10','DC10',
                'CF11','CG11','CH11','CI11','CJ11','CK11','CL11','CM11','CN11','CO11','CP11','CQ11','CR11','CS11','CT11','CU11','CV11','CW11','CX11','CY11','CZ11','DA11','DB11','DC11',
                'CF12','CG12','CH12','CI12','CJ12','CK12','CL12','CM12','CN12','CO12','CP12','CQ12','CR12','CS12','CT12','CU12','CV12','CW12','CX12','CY12','CZ12','DA12','DB12','DC12',
                'CF13','CG13','CH13','CI13','CJ13','CK13','CL13','CM13','CN13','CO13','CP13','CQ13','CR13','CS13','CT13','CU13','CV13','CW13','CX13','CY13','CZ13','DA13','DB13','DC13',]

    # template list for DAE14 in the spreadsheet
    DAE14list = ['CF15','CG15','CH15','CI15','CJ15','CK15','CL15','CM15','CN15','CO15','CP15','CQ15','CR15','CS15','CT15','CU15','CV15','CW15','CX15','CY15','CZ15','DA15','DB15','DC15',
                'CF16','CG16','CH16','CI16','CJ16','CK16','CL16','CM16','CN16','CO16','CP16','CQ16','CR16','CS16','CT16','CU16','CV16','CW16','CX16','CY16','CZ16','DA16','DB16','DC16',
                'CF17','CG17','CH17','CI17','CJ17','CK17','CL17','CM17','CN17','CO17','CP17','CQ17','CR17','CS17','CT17','CU17','CV17','CW17','CX17','CY17','CZ17','DA17','DB17','DC17',
                'CF18','CG18','CH18','CI18','CJ18','CK18','CL18','CM18','CN18','CO18','CP18','CQ18','CR18','CS18','CT18','CU18','CV18','CW18','CX18','CY18','CZ18','DA18','DB18','DC18',
                'CF19','CG19','CH19','CI19','CJ19','CK19','CL19','CM19','CN19','CO19','CP19','CQ19','CR19','CS19','CT19','CU19','CV19','CW19','CX19','CY19','CZ19','DA19','DB19','DC19',]

    # template list for DAE15 in the spreadsheet
    DAE15list = ['CF21','CG21','CH21','CI21','CJ21','CK21','CL21','CM21','CN21','CO21','CP21','CQ21','CR21','CS21','CT21','CU21','CV21','CW21','CX21','CY21','CZ21','DA21','DB21','DC21',
                'CF22','CG22','CH22','CI22','CJ22','CK22','CL22','CM22','CN22','CO22','CP22','CQ22','CR22','CS22','CT22','CU22','CV22','CW22','CX22','CY22','CZ22','DA22','DB22','DC22',
                'CF23','CG23','CH23','CI23','CJ23','CK23','CL23','CM23','CN23','CO23','CP23','CQ23','CR23','CS23','CT23','CU23','CV23','CW23','CX23','CY23','CZ23','DA23','DB23','DC23',
                'CF24','CG24','CH24','CI24','CJ24','CK24','CL24','CM24','CN24','CO24','CP24','CQ24','CR24','CS24','CT24','CU24','CV24','CW24','CX24','CY24','CZ24','DA24','DB24','DC24',
                'CF25','CG25','CH25','CI25','CJ25','CK25','CL25','CM25','CN25','CO25','CP25','CQ25','CR25','CS25','CT25','CU25','CV25','CW25','CX25','CY25','CZ25','DA25','DB25','DC25',]

    # template list for DAE16 in the spreadsheet
    DAE16list = ['CF27','CG27','CH27','CI27','CJ27','CK27','CL27','CM27','CN27','CO27','CP27','CQ27','CR27','CS27','CT27','CU27','CV27','CW27','CX27','CY27','CZ27','DA27','DB27','DC27',
                'CF28','CG28','CH28','CI28','CJ28','CK28','CL28','CM28','CN28','CO28','CP28','CQ28','CR28','CS28','CT28','CU28','CV28','CW28','CX28','CY28','CZ28','DA28','DB28','DC28',
                'CF29','CG29','CH29','CI29','CJ29','CK29','CL29','CM29','CN29','CO29','CP29','CQ29','CR29','CS29','CT29','CU29','CV29','CW29','CX29','CY29','CZ29','DA29','DB29','DC29',
                'CF30','CG30','CH30','CI30','CJ30','CK30','CL30','CM30','CN30','CO30','CP30','CQ30','CR30','CS30','CT30','CU30','CV30','CW30','CX30','CY30','CZ30','DA30','DB30','DC30',
                'CF31','CG31','CH31','CI31','CJ31','CK31','CL31','CM31','CN31','CO31','CP31','CQ31','CR31','CS31','CT31','CU31','CV31','CW31','CX31','CY31','CZ31','DA31','DB31','DC31',]

    ENGcounter = 0
    cacheslots = 0
    while cacheslots < 24:
        for x in range(0, 8):
            #print ENGcounter
            #test_cell = '=IF($C$19+$C$7=8,IF(($C$15+(($C$23/1000)*($C$15/$C$14)))/($C$7+$C$19)/20>' + str(cacheslots + 1) + ',1),IF($C$15+($C$23*($C$15/$C$14))/($C$7+$C$19)/20>' + str(cacheslots + 1) + '1,0))'
            test_cell = '=IF($C$19+$C$7>=1,IF(ROUND((($C$13)*($C$10))/($C$7+$C$19)/20,0)>=' + str(cacheslots + 1) + ',1,0))'
            ws[ENG1list[ENGcounter]] = test_cell
            test_cell = '=IF($C$19+$C$7>=2,IF(ROUND((($C$13)*($C$10))/($C$7+$C$19)/20,0)>=' + str(cacheslots + 1) + ',1,0))'
            ws[ENG2list[ENGcounter]] = test_cell
            test_cell = '=IF($C$19+$C$7>=3,IF(ROUND((($C$13)*($C$10))/($C$7+$C$19)/20,0)>=' + str(cacheslots + 1) + ',1,0))'
            ws[ENG3list[ENGcounter]] = test_cell
            test_cell = '=IF($C$19+$C$7>=4,IF(ROUND((($C$13)*($C$10))/($C$7+$C$19)/20,0)>=' + str(cacheslots + 1) + ',1,0))'
            ws[ENG4list[ENGcounter]] = test_cell
            test_cell = '=IF($C$19+$C$7>=5,IF(ROUND((($C$13)*($C$10))/($C$7+$C$19)/20,0)>=' + str(cacheslots + 1) + ',1,0))'
            ws[ENG5list[ENGcounter]] = test_cell
            test_cell = '=IF($C$19+$C$7>=6,IF(ROUND((($C$13)*($C$10))/($C$7+$C$19)/20,0)>=' + str(cacheslots + 1) + ',1,0))'
            ws[ENG6list[ENGcounter]] = test_cell
            test_cell = '=IF($C$19+$C$7>=7,IF(ROUND((($C$13)*($C$10))/($C$7+$C$19)/20,0)>=' + str(cacheslots + 1) + ',1,0))'
            #test_cell = '=IF($C$19+$C$7=8,IF(($C$15+(($C$23/1000)*($C$15/$C$14)))/($C$7+$C$19)/20>' + str(cacheslots + 1) + ',1,0))'
            ws[ENG7list[ENGcounter]] = test_cell
            test_cell = '=IF($C$19+$C$7>=8,IF(ROUND((($C$13)*($C$10))/($C$7+$C$19)/20,0)>=' + str(cacheslots + 1) + ',1,0))'
            ws[ENG8list[ENGcounter]] = test_cell
        cacheslots = cacheslots + 1
        ENGcounter = ENGcounter + 1

    # keeping track of the DAE slots to make sure we do something with all of them
    DAE1counter = 0
    DAE2counter = 0
    DAE3counter = 0
    DAE4counter = 0
    DAE5counter = 0
    DAE6counter = 0
    DAE7counter = 0
    DAE8counter = 0
    DAE9counter = 0
    DAE10counter = 0
    DAE11counter = 0
    DAE12counter = 0
    DAE13counter = 0
    DAE14counter = 0
    DAE15counter = 0
    DAE16counter = 0

    # formulas that are needed in the spreadsheet
    #   '=IF($C$7=6,IF($C$22>=7,3558,0),IF($C$22>=9,3558,0))',
    #   '=IF($C$7=6,IF($C$22>=13,3558,0),IF($C$22>=17,3558,0))',
    #   '=IF($C$7=6,IF($C$22>=19,3558,0),IF($C$22>=25,3558,0))',
        
    #fobj = zopen(zip_fileobj, xml_file)
    #print f
    # read the xml file into memory
    if f != None:
        tree = ET.parse(f)
        root = tree.getroot()

    # counting total disk in the array
    disk_counter = 0

    test_cell = ""

    # iterating thru xml
    for child in root.iter():
        #print child
        #print 'tag', child.tag
        #print 'text', child.text

        symid = child.find('symid')
        if symid != None:
            s = str(symid.text)
            serialnumber = s.strip('0')
            ws['B4'] = 'VMAX ' + serialnumber

        # searching for ident which is the VMAX director and DAE
        ident = child.find('ident')
        if ident != None:
            #print test_cell
            #print ident.text
            # fill the DAE with the number of disk drives found
            if ident.text == "DF-1C":  
            #interface = child.find('interface')
            #if interface != None:
            #tid = child.find('tid')
            #   if interface != None:
                megabytes = child.find('actual_megabytes')
                failed = child.find('failed_disk')
                #print failed.text
                # converting text to integer
                i = int(megabytes.text)
                i = i/1024
                ws[DAE1list[DAE1counter]] = i
                DAE1counter = DAE1counter + 1
                disk_counter = disk_counter + 1
            else:
                #if lastDAE == "DF-1C":
                if DAECounter == 1:
                    print 'there are ' + str(DAE1counter) + ' disks in DAE1'
                    # fill the rest of the DAE before going to the next one
                    remainingslots = remainingslots - DAE1counter
                    while remainingslots > 4:
                        for x in range(0, 4):
                            #print "We're on time %d" % (x)
                            if engine_counter == 6:
                                test_cell = '=IF(($C$7+$C$19=' + str(EngineCount) + '),IF($C$22>=' + str(BalanceTracker + 1) + ',3558,0),IF($C$22>=' + str(BalanceTracker + 3) + ',3558,0))'
                            if engine_counter == 8:
                                test_cell = '=IF($C$22>=' + str(BalanceTracker + 1) + ',3558,0)'
                            ws[DAE1list[DAE1counter]] = test_cell
                            DAE1counter = DAE1counter + 1
                        BalanceTracker = BalanceTracker + (DAECount / 2)
                        remainingslots = remainingslots - 4
                    BalanceTracker = 0
                    remainingslots = 119
                    DAECounter = DAECounter + 1
                #print ident.text, interface.text, tid.text, megabytes.text
                

            
            if ident.text == "DF-2C":  
            #interface = child.find('interface')
            #if interface != None:
            #tid = child.find('tid')
            #   if interface != None:
                megabytes = child.find('actual_megabytes')
                failed = child.find('failed_disk')
                #print failed.text
                # converting text to integer
                i = int(megabytes.text)
                i = i/1024
                # Data can be assigned directly to cells
                ws[DAE2list[DAE2counter]] = i
                DAE2counter = DAE2counter + 1
                disk_counter = disk_counter + 1
            else:
                #if lastDAE == "DF-2C":
                if DAECounter == 2:
                    print 'there are ' + str(DAE2counter) + ' disks in DAE2'
                    # fill the rest of the DAE before going to the next one
                    remainingslots = remainingslots - DAE2counter
                    while remainingslots > 4:
                        for x in range(0, 4):
                            #print "We're on time %d" % (x)
                            if engine_counter == 6:
                                test_cell = '=IF(($C$7+$C$19=' + str(EngineCount) + '),IF($C$22>=' + str(BalanceTracker + 1) + ',3558,0),IF($C$22>=' + str(BalanceTracker + 3) + ',3558,0))'
                            if engine_counter == 8:
                                test_cell = '=IF($C$22>=' + str(BalanceTracker + 1) + ',3558,0)'
                            ws[DAE2list[DAE2counter]] = test_cell
                            DAE2counter = DAE2counter + 1
                        BalanceTracker = BalanceTracker + (DAECount / 2)
                        remainingslots = remainingslots - 4
                    BalanceTracker = 0
                    remainingslots = 119
                    DAECounter = DAECounter + 1
                #print ident.text, interface.text, tid.text, megabytes.text
                

            if ident.text == "DF-3C":  
            #interface = child.find('interface')
            #if interface != None:
            #tid = child.find('tid')
            #   if interface != None:
                megabytes = child.find('actual_megabytes')
                failed = child.find('failed_disk')
                #print failed.text

                # converting text to integer
                i = int(megabytes.text)
                i = i/1024
                # Data can be assigned directly to cells
                ws[DAE3list[DAE3counter]] = i
                DAE3counter = DAE3counter + 1
                disk_counter = disk_counter + 1
            else:
                #if lastDAE == "DF-3C":
                if DAECounter == 3:
                    print 'there are ' + str(DAE3counter) + ' disks in DAE3'
                    # fill the rest of the DAE before going to the next one
                    remainingslots = remainingslots - DAE3counter
                    while remainingslots > 4:
                        for x in range(0, 4):
                            #print "We're on time %d" % (x)
                            if engine_counter == 6:
                                test_cell = '=IF(($C$7+$C$19=' + str(EngineCount) + '),IF($C$22>=' + str(BalanceTracker + 2) + ',3558,0),IF($C$22>=' + str(BalanceTracker + 4) + ',3558,0))'
                            if engine_counter == 8:
                                test_cell = '=IF($C$22>=' + str(BalanceTracker + 2) + ',3558,0)'
                            ws[DAE3list[DAE3counter]] = test_cell
                            DAE3counter = DAE3counter + 1
                        BalanceTracker = BalanceTracker + (DAECount / 2)
                        remainingslots = remainingslots - 4
                    BalanceTracker = 0
                    remainingslots = 119
                    DAECounter = DAECounter + 1
                #print ident.text, interface.text, tid.text, megabytes.text


            if ident.text == "DF-4C":  
            #interface = child.find('interface')
            #if interface != None:
            #tid = child.find('tid')
            #   if interface != None:
                megabytes = child.find('actual_megabytes')
                failed = child.find('failed_disk')
                #print failed.text

                # converting text to integer
                i = int(megabytes.text)
                i = i/1024
                # Data can be assigned directly to cells
                ws[DAE4list[DAE4counter]] = i
                DAE4counter = DAE4counter + 1
                disk_counter = disk_counter + 1
            else:
                #if lastDAE == "DF-4C":
                if DAECounter == 4:
                    print 'there are ' + str(DAE4counter) + ' disks in DAE4'
                    # fill the rest of the DAE before going to the next one
                    remainingslots = remainingslots - DAE4counter
                    while remainingslots > 4:
                        for x in range(0, 4):
                            #print "We're on time %d" % (x)
                            if engine_counter == 6:
                                test_cell = '=IF(($C$7+$C$19=' + str(EngineCount) + '),IF($C$22>=' + str(BalanceTracker + 2) + ',3558,0),IF($C$22>=' + str(BalanceTracker + 4) + ',3558,0))'
                            if engine_counter == 8:
                                test_cell = '=IF($C$22>=' + str(BalanceTracker + 2) + ',3558,0)'
                            ws[DAE4list[DAE4counter]] = test_cell
                            DAE4counter = DAE4counter + 1
                        BalanceTracker = BalanceTracker + (DAECount / 2)
                        remainingslots = remainingslots - 4
                    BalanceTracker = 0
                    remainingslots = 119
                    DAECounter = DAECounter + 1
                #print ident.text, interface.text, tid.text, megabytes.text


            if ident.text == "DF-5C":  
            #interface = child.find('interface')
            #if interface != None:
            #tid = child.find('tid')
            #   if interface != None:
                megabytes = child.find('actual_megabytes')
                failed = child.find('failed_disk')
                #print failed.text

                # converting text to integer
                i = int(megabytes.text)
                i = i/1024
                # Data can be assigned directly to cells
                ws[DAE5list[DAE5counter]] = i
                DAE5counter = DAE5counter + 1
                disk_counter = disk_counter + 1
            else:
                #if lastDAE == "DF-5C":
                if DAECounter == 5:
                    print 'there are ' + str(DAE5counter) + ' disks in DAE5'
                    # fill the rest of the DAE before going to the next one
                    remainingslots = remainingslots - DAE5counter
                    while remainingslots > 4:
                        for x in range(0, 4):
                            #print "We're on time %d" % (x)
                            if engine_counter == 6:
                                test_cell = '=IF(($C$7+$C$19=' + str(EngineCount) + '),IF($C$22>=' + str(BalanceTracker + 3) + ',3558,0),IF($C$22>=' + str(BalanceTracker + 5) + ',3558,0))'
                            if engine_counter == 8:
                                test_cell = '=IF($C$22>=' + str(BalanceTracker + 3) + ',3558,0)'
                            ws[DAE5list[DAE5counter]] = test_cell
                            DAE5counter = DAE5counter + 1
                        BalanceTracker = BalanceTracker + (DAECount / 2)
                        remainingslots = remainingslots - 4
                    BalanceTracker = 0
                    remainingslots = 119
                    DAECounter = DAECounter + 1
                #print ident.text, interface.text, tid.text, megabytes.text


            if ident.text == "DF-6C":  
            #interface = child.find('interface')
            #if interface != None:
            #tid = child.find('tid')
            #   if interface != None:
                megabytes = child.find('actual_megabytes')
                failed = child.find('failed_disk')
                #print failed.text

                # converting text to integer
                i = int(megabytes.text)
                i = i/1024
                # Data can be assigned directly to cells
                ws[DAE6list[DAE6counter]] = i
                DAE6counter = DAE6counter + 1
                disk_counter = disk_counter + 1
            else:
                #if lastDAE == "DF-6C":
                if DAECounter == 6:
                    print 'there are ' + str(DAE6counter) + ' disks in DAE6'
                    # fill the rest of the DAE before going to the next one
                    remainingslots = remainingslots - DAE6counter
                    while remainingslots > 4:
                        for x in range(0, 4):
                            #print "We're on time %d" % (x)
                            if engine_counter == 6:
                                test_cell = '=IF(($C$7+$C$19=' + str(EngineCount) + '),IF($C$22>=' + str(BalanceTracker + 3) + ',3558,0),IF($C$22>=' + str(BalanceTracker + 5) + ',3558,0))'
                            if engine_counter == 8:
                                test_cell = '=IF($C$22>=' + str(BalanceTracker + 3) + ',3558,0)'
                            ws[DAE6list[DAE6counter]] = test_cell
                            DAE6counter = DAE6counter + 1
                        BalanceTracker = BalanceTracker + (DAECount / 2)
                        remainingslots = remainingslots - 4
                    BalanceTracker = 0
                    remainingslots = 119
                    DAECounter = DAECounter + 1
                #print ident.text, interface.text, tid.text, megabytes.text


            if ident.text == "DF-7C":  
            #interface = child.find('interface')
            #if interface != None:
            #tid = child.find('tid')
            #   if interface != None:
                megabytes = child.find('actual_megabytes')
                failed = child.find('failed_disk')
                #print failed.text

                # converting text to integer
                i = int(megabytes.text)
                i = i/1024
                # Data can be assigned directly to cells
                ws[DAE7list[DAE7counter]] = i
                DAE7counter = DAE7counter + 1
                disk_counter = disk_counter + 1
            else:
                #if lastDAE == "DF-7C":
                if DAECounter == 7:
                    print 'there are ' + str(DAE7counter) + ' disks in DAE7'
                    # fill the rest of the DAE before going to the next one
                    remainingslots = remainingslots - DAE7counter
                    while remainingslots > 4:
                        for x in range(0, 4):
                            #print "We're on time %d" % (x)
                            if engine_counter == 6:
                                test_cell = '=IF(($C$7+$C$19=' + str(EngineCount) + '),IF($C$22>=' + str(BalanceTracker + 4) + ',3558,0),IF($C$22>=' + str(BalanceTracker + 6) + ',3558,0))'
                            if engine_counter == 8:
                                test_cell = '=IF($C$22>=' + str(BalanceTracker + 4) + ',3558,0)'
                            ws[DAE7list[DAE7counter]] = test_cell
                            DAE7counter = DAE7counter + 1
                        BalanceTracker = BalanceTracker + (DAECount / 2)
                        remainingslots = remainingslots - 4
                    BalanceTracker = 0
                    remainingslots = 119
                    DAECounter = DAECounter + 1
                #print ident.text, interface.text, tid.text, megabytes.text
                
            if ident.text == "DF-8C":  
            #interface = child.find('interface')
            #if interface != None:
            #tid = child.find('tid')
            #   if interface != None:
                megabytes = child.find('actual_megabytes')
                failed = child.find('failed_disk')
                #print failed.text

                # converting text to integer
                i = int(megabytes.text)
                i = i/1024
                # Data can be assigned directly to cells
                ws[DAE8list[DAE8counter]] = i
                DAE8counter = DAE8counter + 1
                disk_counter = disk_counter + 1
            else:
                #if lastDAE == "DF-8C":
                if DAECounter == 8:
                    print 'there are ' + str(DAE8counter) + ' disks in DAE8'
                    # fill the rest of the DAE before going to the next one
                    remainingslots = remainingslots - DAE8counter
                    while remainingslots > 4:
                        for x in range(0, 4):
                            #print "We're on time %d" % (x)
                            if engine_counter == 6:
                                test_cell = '=IF(($C$7+$C$19=' + str(EngineCount) + '),IF($C$22>=' + str(BalanceTracker + 4) + ',3558,0),IF($C$22>=' + str(BalanceTracker + 6) + ',3558,0))'
                            if engine_counter == 8:
                                test_cell = '=IF($C$22>=' + str(BalanceTracker + 4) + ',3558,0)'
                            ws[DAE8list[DAE8counter]] = test_cell
                            DAE8counter = DAE8counter + 1
                        BalanceTracker = BalanceTracker + (DAECount / 2)
                        remainingslots = remainingslots - 4
                    BalanceTracker = 0
                    remainingslots = 119
                    DAECounter = DAECounter + 1
                #print ident.text, interface.text, tid.text, megabytes.text



            if ident.text == "DF-9C":  
            #interface = child.find('interface')
            #if interface != None:
            #tid = child.find('tid')
            #   if interface != None:
                megabytes = child.find('actual_megabytes')
                failed = child.find('failed_disk')
                #print failed.text

                # converting text to integer
                i = int(megabytes.text)
                i = i/1024
                # Data can be assigned directly to cells
                ws[DAE9list[DAE9counter]] = i
                DAE9counter = DAE9counter + 1
                disk_counter = disk_counter + 1
            else:
                #if lastDAE == "DF-9C":
                if DAECounter == 9:
                    print 'there are ' + str(DAE9counter) + ' disks in DAE9'
                    # fill the rest of the DAE before going to the next one
                    remainingslots = remainingslots - DAE9counter
                    while remainingslots > 4:
                        for x in range(0, 4):
                            #print "We're on time %d" % (x)
                            if engine_counter == 6:
                                test_cell = '=IF(($C$7+$C$19=' + str(EngineCount) + '),IF($C$22>=' + str(BalanceTracker + 5) + ',3558,0),IF($C$22>=' + str(BalanceTracker + 7) + ',3558,0))'
                            if engine_counter == 8:
                                test_cell = '=IF($C$22>=' + str(BalanceTracker + 5) + ',3558,0)'
                            ws[DAE9list[DAE9counter]] = test_cell
                            DAE9counter = DAE9counter + 1
                        BalanceTracker = BalanceTracker + (DAECount / 2)
                        remainingslots = remainingslots - 4
                    BalanceTracker = 0
                    remainingslots = 119
                    DAECounter = DAECounter + 1
                #print ident.text, interface.text, tid.text, megabytes.text

            if ident.text == "DF-10C":  
            #interface = child.find('interface')
            #if interface != None:
            #tid = child.find('tid')
            #   if interface != None:
                megabytes = child.find('actual_megabytes')
                failed = child.find('failed_disk')
                #print failed.text

                # converting text to integer
                i = int(megabytes.text)
                i = i/1024
                # Data can be assigned directly to cells
                ws[DAE10list[DAE10counter]] = i
                DAE10counter = DAE10counter + 1
                disk_counter = disk_counter + 1
            else:
                #if lastDAE == "DF-10C":
                if DAECounter == 10:
                    print 'there are ' + str(DAE10counter) + ' disks in DAE10'
                    # fill the rest of the DAE before going to the next one
                    remainingslots = remainingslots - DAE10counter
                    while remainingslots > 4:
                        for x in range(0, 4):
                            #print "We're on time %d" % (x)
                            if engine_counter == 6:
                                test_cell = '=IF(($C$7+$C$19=' + str(EngineCount) + '),IF($C$22>=' + str(BalanceTracker + 5) + ',3558,0),IF($C$22>=' + str(BalanceTracker + 7) + ',3558,0))'
                            if engine_counter == 8:
                                test_cell = '=IF($C$22>=' + str(BalanceTracker + 5) + ',3558,0)'
                            ws[DAE10list[DAE10counter]] = test_cell
                            DAE10counter = DAE10counter + 1
                        BalanceTracker = BalanceTracker + (DAECount / 2)
                        remainingslots = remainingslots - 4
                    BalanceTracker = 0
                    remainingslots = 119
                    DAECounter = DAECounter + 1
                #print ident.text, interface.text, tid.text, megabytes.text


            if ident.text == "DF-11C":  
            #interface = child.find('interface')
            #if interface != None:
            #tid = child.find('tid')
            #   if interface != None:
                megabytes = child.find('actual_megabytes')
                failed = child.find('failed_disk')
                #print failed.text

                # converting text to integer
                i = int(megabytes.text)
                i = i/1024
                # Data can be assigned directly to cells
                ws[DAE11list[DAE11counter]] = i
                DAE11counter = DAE11counter + 1
                disk_counter = disk_counter + 1
            else:
                #if lastDAE == "DF-11C":
                if DAECounter == 11:
                    print 'there are ' + str(DAE11counter) + ' disks in DAE11'
                    # fill the rest of the DAE before going to the next one
                    remainingslots = remainingslots - DAE11counter
                    while remainingslots > 4:
                        for x in range(0, 4):
                            #print "We're on time %d" % (x)
                            if engine_counter == 6:
                                test_cell = '=IF(($C$7+$C$19=' + str(EngineCount) + '),IF($C$22>=' + str(BalanceTracker + 6) + ',3558,0),IF($C$22>=' + str(BalanceTracker + 8) + ',3558,0))'
                            if engine_counter == 8:
                                test_cell = '=IF($C$22>=' + str(BalanceTracker + 6) + ',3558,0)'
                            ws[DAE11list[DAE11counter]] = test_cell
                            DAE11counter = DAE11counter + 1
                        BalanceTracker = BalanceTracker + (DAECount / 2)
                        remainingslots = remainingslots - 4
                    BalanceTracker = 0
                    remainingslots = 119
                    DAECounter = DAECounter + 1
                #print ident.text, interface.text, tid.text, megabytes.text


            if ident.text == "DF-12C":  
            #interface = child.find('interface')
            #if interface != None:
            #tid = child.find('tid')
            #   if interface != None:
                megabytes = child.find('actual_megabytes')
                failed = child.find('failed_disk')
                #print failed.text

                # converting text to integer
                i = int(megabytes.text)
                i = i/1024
                # Data can be assigned directly to cells
                ws[DAE12list[DAE12counter]] = i
                DAE12counter = DAE12counter + 1
                disk_counter = disk_counter + 1
            else:
                #if lastDAE == "DF-12C":
                if DAECounter == 12:
                    print 'there are ' + str(DAE12counter) + ' disks in DAE12'
                    # fill the rest of the DAE before going to the next one
                    remainingslots = remainingslots - DAE12counter
                    while remainingslots > 4:
                        for x in range(0, 4):
                            #print "We're on time %d" % (x)
                            if engine_counter == 6:
                                test_cell = '=IF(($C$7+$C$19=' + str(EngineCount) + '),IF($C$22>=' + str(BalanceTracker + 6) + ',3558,0),IF($C$22>=' + str(BalanceTracker + 8) + ',3558,0))'
                            if engine_counter == 8:
                                test_cell = '=IF($C$22>=' + str(BalanceTracker + 6) + ',3558,0)'
                            ws[DAE12list[DAE12counter]] = test_cell
                            DAE12counter = DAE12counter + 1
                        BalanceTracker = BalanceTracker + (DAECount / 2)
                        remainingslots = remainingslots - 4
                    BalanceTracker = 0
                    remainingslots = 119
                    DAECounter = DAECounter + 1
                #print ident.text, interface.text, tid.text, megabytes.text


            if ident.text == "DF-13C":  
            #interface = child.find('interface')
            #if interface != None:
            #tid = child.find('tid')
            #   if interface != None:
                megabytes = child.find('actual_megabytes')
                failed = child.find('failed_disk')
                #print failed.text

                # converting text to integer
                i = int(megabytes.text)
                i = i/1024
                # Data can be assigned directly to cells
                ws[DAE13list[DAE13counter]] = i
                DAE13counter = DAE13counter + 1
                disk_counter = disk_counter + 1
            else:
                #if lastDAE == "DF-13C":
                if DAECounter == 13:
                    print 'there are ' + str(DAE13counter) + ' disks in DAE13'
                    # fill the rest of the DAE before going to the next one
                    remainingslots = remainingslots - DAE13counter
                    while remainingslots > 4:
                        for x in range(0, 4):
                            #print "We're on time %d" % (x)
                            if engine_counter == 6:
                                test_cell = '=IF($C$19=0,0,IF($C$22>=' + str(BalanceTracker + 7) + ',3558,0))'
                            if engine_counter == 8:
                                test_cell = '=IF($C$22>=' + str(BalanceTracker + 7) + ',3558,0)'
                            ws[DAE13list[DAE13counter]] = test_cell
                            DAE13counter = DAE13counter + 1
                        BalanceTracker = BalanceTracker + (DAECount / 2)
                        remainingslots = remainingslots - 4
                    BalanceTracker = 0
                    remainingslots = 119
                    DAECounter = DAECounter + 1
                #print ident.text, interface.text, tid.text, megabytes.text


            if ident.text == "DF-14C":  
            #interface = child.find('interface')
            #if interface != None:
            #tid = child.find('tid')
            #   if interface != None:
                megabytes = child.find('actual_megabytes')
                failed = child.find('failed_disk')
                #print failed.text

                # converting text to integer
                i = int(megabytes.text)
                i = i/1024
                # Data can be assigned directly to cells
                ws[DAE14list[DAE14counter]] = i
                DAE14counter = DAE14counter + 1
                disk_counter = disk_counter + 1
            else:
                #if lastDAE == "DF-14C":
                if DAECounter == 14:
                    print 'there are ' + str(DAE14counter) + ' disks in DAE14'
                    # fill the rest of the DAE before going to the next one
                    remainingslots = remainingslots - DAE14counter
                    while remainingslots > 4:
                        for x in range(0, 4):
                            #print "We're on time %d" % (x)
                            if engine_counter == 6:
                                test_cell = '=IF($C$19=0,0,IF($C$22>=' + str(BalanceTracker + 7) + ',3558,0))'
                            if engine_counter == 8:
                                test_cell = '=IF($C$22>=' + str(BalanceTracker + 7) + ',3558,0)'
                            ws[DAE14list[DAE14counter]] = test_cell
                            DAE14counter = DAE14counter + 1
                        BalanceTracker = BalanceTracker + (DAECount / 2)
                        remainingslots = remainingslots - 4
                    BalanceTracker = 0
                    remainingslots = 119
                    DAECounter = DAECounter + 1
                #print ident.text, interface.text, tid.text, megabytes.text


            if ident.text == "DF-15C":  
            #interface = child.find('interface')
            #if interface != None:
            #tid = child.find('tid')
            #   if interface != None:
                megabytes = child.find('actual_megabytes')
                failed = child.find('failed_disk')
                #print failed.text

                # converting text to integer
                i = int(megabytes.text)
                i = i/1024
                # Data can be assigned directly to cells
                ws[DAE15list[DAE15counter]] = i
                DAE15counter = DAE15counter + 1
                disk_counter = disk_counter + 1
            else:
                #if lastDAE == "DF-15C":
                if DAECounter == 15:
                    print 'there are ' + str(DAE15counter) + ' disks in DAE15'
                    # fill the rest of the DAE before going to the next one
                    remainingslots = remainingslots - DAE15counter
                    while remainingslots > 4:
                        for x in range(0, 4):
                            #print "We're on time %d" % (x)
                            if engine_counter == 6:
                                test_cell = '=IF($C$19=0,0,IF($C$22>=' + str(BalanceTracker + 7) + ',3558,0))'
                            if engine_counter == 8:
                                test_cell = '=IF($C$22>=' + str(BalanceTracker + 8) + ',3558,0)'
                            ws[DAE15list[DAE15counter]] = test_cell
                            DAE15counter = DAE15counter + 1
                        BalanceTracker = BalanceTracker + (DAECount / 2)
                        remainingslots = remainingslots - 4
                    BalanceTracker = 0
                    remainingslots = 119
                    DAECounter = DAECounter + 1
                #print ident.text, interface.text, tid.text, megabytes.text


            if ident.text == "DF-16C":  
            #interface = child.find('interface')
            #if interface != None:
            #tid = child.find('tid')
            #   if interface != None:
                megabytes = child.find('actual_megabytes')
                failed = child.find('failed_disk')
                #print failed.text

                # converting text to integer
                i = int(megabytes.text)
                i = i/1024
                # Data can be assigned directly to cells
                ws[DAE16list[DAE16counter]] = i
                DAE16counter = DAE16counter + 1
                disk_counter = disk_counter + 1
            else:
                #if lastDAE == "DF-16C":
                if DAECounter == 16:
                    print 'there are ' + str(DAE16counter) + ' disks in DAE16'
                    # fill the rest of the DAE before going to the next one
                    remainingslots = remainingslots - DAE16counter
                    while remainingslots > 4:
                        for x in range(0, 4):
                            #print "We're on time %d" % (x)
                            if engine_counter == 6:
                                test_cell = '=IF($C$19=0,0,IF($C$22>=' + str(BalanceTracker + 7) + ',3558,0))'
                            if engine_counter == 8:
                                test_cell = '=IF($C$22>=' + str(BalanceTracker + 8) + ',3558,0)'
                            ws[DAE16list[DAE16counter]] = test_cell
                            DAE16counter = DAE16counter + 1
                        BalanceTracker = BalanceTracker + (DAECount / 2)
                        remainingslots = remainingslots - 4
                    BalanceTracker = 0
                    remainingslots = 119
                    DAECounter = DAECounter + 1
                #print ident.text, interface.text, tid.text, megabytes.text

            lastDAE = ident.text
            #print str(DAECounter) + ' DAE counter'



    while DAECounter < 16:
        #print 'DAECounter ' + str(DAECounter)
        if DAECounter == 1:
            if DAE1counter != 0:
                DAEnumber = 1
            print 'there are ' + str(DAE1counter) + ' disks in DAE1'
            # fill the rest of the DAE before going to the next one
            remainingslots = remainingslots - DAE1counter
            while remainingslots > 4:
                for x in range(0, 4):
                    #print "We're on time %d" % (x)
                    if engine_counter == 6:
                        test_cell = '=IF(($C$7+$C$19=' + str(EngineCount) + '),IF($C$22>=' + str(BalanceTracker + 1) + ',3558,0),IF($C$22>=' + str(BalanceTracker + 3) + ',3558,0))'
                    if engine_counter == 8:
                        test_cell = '=IF($C$22>=' + str(BalanceTracker + 1) + ',3558,0)'
                    ws[DAE1list[DAE1counter]] = test_cell
                    DAE1counter = DAE1counter + 1
                BalanceTracker = BalanceTracker + (DAECount / 2)
                remainingslots = remainingslots - 4
            BalanceTracker = 0
            remainingslots = 119
            DAECounter = DAECounter + 1
                #print ident.text, interface.text, tid.text, megabytes.text

        if DAECounter == 2:
            if DAE2counter != 0:
                DAEnumber = 2
            print 'there are ' + str(DAE2counter) + ' disks in DAE2'
            # fill the rest of the DAE before going to the next one
            remainingslots = remainingslots - DAE2counter
            while remainingslots > 4:
                for x in range(0, 4):
                    #print "We're on time %d" % (x)
                    if engine_counter == 6:
                        test_cell = '=IF(($C$7+$C$19=' + str(EngineCount) + '),IF($C$22>=' + str(BalanceTracker + 1) + ',3558,0),IF($C$22>=' + str(BalanceTracker + 3) + ',3558,0))'
                    if engine_counter == 8:
                        test_cell = '=IF($C$22>=' + str(BalanceTracker + 1) + ',3558,0)'
                    ws[DAE2list[DAE2counter]] = test_cell
                    DAE2counter = DAE2counter + 1
                BalanceTracker = BalanceTracker + (DAECount / 2)
                remainingslots = remainingslots - 4
            BalanceTracker = 0
            remainingslots = 119
            DAECounter = DAECounter + 1
                #print ident.text, interface.text, tid.text, megabytes.text

        if DAECounter == 3:
            if DAE3counter != 0:
                DAEnumber = 3
            print 'there are ' + str(DAE3counter) + ' disks in DAE3'
            # fill the rest of the DAE before going to the next one
            remainingslots = remainingslots - DAE3counter
            while remainingslots > 4:
                for x in range(0, 4):
                    #print "We're on time %d" % (x)
                    if engine_counter == 6:
                        test_cell = '=IF(($C$7+$C$19=' + str(EngineCount) + '),IF($C$22>=' + str(BalanceTracker + 2) + ',3558,0),IF($C$22>=' + str(BalanceTracker + 4) + ',3558,0))'
                    if engine_counter == 8:
                        test_cell = '=IF($C$22>=' + str(BalanceTracker + 2) + ',3558,0)'
                    ws[DAE3list[DAE3counter]] = test_cell
                    DAE3counter = DAE3counter + 1
                BalanceTracker = BalanceTracker + (DAECount / 2)
                remainingslots = remainingslots - 4
            BalanceTracker = 0
            remainingslots = 119
            DAECounter = DAECounter + 1
                #print ident.text, interface.text, tid.text, megabytes.text

        if DAECounter == 4:
            if DAE4counter != 0:
                DAEnumber = 4
            print 'there are ' + str(DAE4counter) + ' disks in DAE4'
            # fill the rest of the DAE before going to the next one
            remainingslots = remainingslots - DAE4counter
            while remainingslots > 4:
                for x in range(0, 4):
                    #print "We're on time %d" % (x)
                    if engine_counter == 6:
                        test_cell = '=IF(($C$7+$C$19=' + str(EngineCount) + '),IF($C$22>=' + str(BalanceTracker + 2) + ',3558,0),IF($C$22>=' + str(BalanceTracker + 4) + ',3558,0))'
                    if engine_counter == 8:
                        test_cell = '=IF($C$22>=' + str(BalanceTracker + 2) + ',3558,0)'
                    ws[DAE4list[DAE4counter]] = test_cell
                    DAE4counter = DAE4counter + 1
                BalanceTracker = BalanceTracker + (DAECount / 2)
                remainingslots = remainingslots - 4
            BalanceTracker = 0
            remainingslots = 119
            DAECounter = DAECounter + 1
                #print ident.text, interface.text, tid.text, megabytes.text

        if DAECounter == 5:
            if DAE5counter != 0:
                DAEnumber = 5
            print 'there are ' + str(DAE5counter) + ' disks in DAE5'
            # fill the rest of the DAE before going to the next one
            remainingslots = remainingslots - DAE5counter
            while remainingslots > 4:
                for x in range(0, 4):
                    #print "We're on time %d" % (x)
                    if engine_counter == 6:
                        test_cell = '=IF(($C$7+$C$19=' + str(EngineCount) + '),IF($C$22>=' + str(BalanceTracker + 3) + ',3558,0),IF($C$22>=' + str(BalanceTracker + 5) + ',3558,0))'
                    if engine_counter == 8:
                        test_cell = '=IF($C$22>=' + str(BalanceTracker + 3) + ',3558,0)'
                    ws[DAE5list[DAE5counter]] = test_cell
                    DAE5counter = DAE5counter + 1
                BalanceTracker = BalanceTracker + (DAECount / 2)
                remainingslots = remainingslots - 4
            BalanceTracker = 0
            remainingslots = 119
            DAECounter = DAECounter + 1
                #print ident.text, interface.text, tid.text, megabytes.text

        if DAECounter == 6:
            if DAE6counter != 0:
                DAEnumber = 6
            print 'there are ' + str(DAE6counter) + ' disks in DAE6'
            # fill the rest of the DAE before going to the next one
            remainingslots = remainingslots - DAE6counter
            while remainingslots > 4:
                for x in range(0, 4):
                    #print "We're on time %d" % (x)
                    if engine_counter == 6:
                        test_cell = '=IF(($C$7+$C$19=' + str(EngineCount) + '),IF($C$22>=' + str(BalanceTracker + 3) + ',3558,0),IF($C$22>=' + str(BalanceTracker + 5) + ',3558,0))'
                    if engine_counter == 8:
                        test_cell = '=IF($C$22>=' + str(BalanceTracker + 3) + ',3558,0)'
                    ws[DAE6list[DAE6counter]] = test_cell
                    DAE6counter = DAE6counter + 1
                BalanceTracker = BalanceTracker + (DAECount / 2)
                remainingslots = remainingslots - 4
            BalanceTracker = 0
            remainingslots = 119
            DAECounter = DAECounter + 1
                #print ident.text, interface.text, tid.text, megabytes.text

        if DAECounter == 7:
            if DAE7counter != 0:
                DAEnumber = 7
            print 'there are ' + str(DAE7counter) + ' disks in DAE7'
            # fill the rest of the DAE before going to the next one
            remainingslots = remainingslots - DAE7counter
            while remainingslots > 4:
                for x in range(0, 4):
                    #print "We're on time %d" % (x)
                    if engine_counter == 6:
                        test_cell = '=IF(($C$7+$C$19=' + str(EngineCount) + '),IF($C$22>=' + str(BalanceTracker + 4) + ',3558,0),IF($C$22>=' + str(BalanceTracker + 6) + ',3558,0))'
                    if engine_counter == 8:
                        test_cell = '=IF($C$22>=' + str(BalanceTracker + 4) + ',3558,0)'
                    ws[DAE7list[DAE7counter]] = test_cell
                    DAE7counter = DAE7counter + 1
                BalanceTracker = BalanceTracker + (DAECount / 2)
                remainingslots = remainingslots - 4
            BalanceTracker = 0
            remainingslots = 119
            DAECounter = DAECounter + 1
                #print ident.text, interface.text, tid.text, megabytes.text

        if DAECounter == 8:
            if DAE8counter != 0:
                DAEnumber = 8
            print 'there are ' + str(DAE8counter) + ' disks in DAE8'
            # fill the rest of the DAE before going to the next one
            remainingslots = remainingslots - DAE8counter
            while remainingslots > 4:
                for x in range(0, 4):
                    #print "We're on time %d" % (x)
                    if engine_counter == 6:
                        test_cell = '=IF(($C$7+$C$19=' + str(EngineCount) + '),IF($C$22>=' + str(BalanceTracker + 4) + ',3558,0),IF($C$22>=' + str(BalanceTracker + 6) + ',3558,0))'
                    if engine_counter == 8:
                        test_cell = '=IF($C$22>=' + str(BalanceTracker + 4) + ',3558,0)'
                    ws[DAE8list[DAE8counter]] = test_cell
                    DAE8counter = DAE8counter + 1
                BalanceTracker = BalanceTracker + (DAECount / 2)
                remainingslots = remainingslots - 4
            BalanceTracker = 0
            remainingslots = 119
            DAECounter = DAECounter + 1
                #print ident.text, interface.text, tid.text, megabytes.text

        if DAECounter == 9:
            if DAE9counter != 0:
                DAEnumber = 9
            print 'there are ' + str(DAE9counter) + ' disks in DAE9'
            # fill the rest of the DAE before going to the next one
            remainingslots = remainingslots - DAE9counter
            while remainingslots > 4:
                for x in range(0, 4):
                    #print "We're on time %d" % (x)
                    if engine_counter == 6:
                        test_cell = '=IF(($C$7+$C$19=' + str(EngineCount) + '),IF($C$22>=' + str(BalanceTracker + 5) + ',3558,0),IF($C$22>=' + str(BalanceTracker + 7) + ',3558,0))'
                    if engine_counter == 8:
                        test_cell = '=IF($C$22>=' + str(BalanceTracker + 5) + ',3558,0)'
                    ws[DAE9list[DAE9counter]] = test_cell
                    DAE9counter = DAE9counter + 1
                BalanceTracker = BalanceTracker + (DAECount / 2)
                remainingslots = remainingslots - 4
            BalanceTracker = 0
            remainingslots = 119
            DAECounter = DAECounter + 1
                #print ident.text, interface.text, tid.text, megabytes.text

        if DAECounter == 10:
            if DAE10counter != 0:
                DAEnumber = 10
            print 'there are ' + str(DAE10counter) + ' disks in DAE10'
            # fill the rest of the DAE before going to the next one
            remainingslots = remainingslots - DAE10counter
            while remainingslots > 4:
                for x in range(0, 4):
                    #print "We're on time %d" % (x)
                    if engine_counter == 6:
                        test_cell = '=IF(($C$7+$C$19=' + str(EngineCount) + '),IF($C$22>=' + str(BalanceTracker + 5) + ',3558,0),IF($C$22>=' + str(BalanceTracker + 7) + ',3558,0))'
                    if engine_counter == 8:
                        test_cell = '=IF($C$22>=' + str(BalanceTracker + 5) + ',3558,0)'
                    ws[DAE10list[DAE10counter]] = test_cell
                    DAE10counter = DAE10counter + 1
                BalanceTracker = BalanceTracker + (DAECount / 2)
                remainingslots = remainingslots - 4
            BalanceTracker = 0
            remainingslots = 119
            DAECounter = DAECounter + 1
                #print ident.text, interface.text, tid.text, megabytes.text

        if DAECounter == 11:
            if DAE11counter != 0:
                DAEnumber = 11
            print 'there are ' + str(DAE11counter) + ' disks in DAE11'
            # fill the rest of the DAE before going to the next one
            remainingslots = remainingslots - DAE11counter
            while remainingslots > 4:
                for x in range(0, 4):
                    #print "We're on time %d" % (x)
                    if engine_counter == 6:
                        test_cell = '=IF(($C$7+$C$19=' + str(EngineCount) + '),IF($C$22>=' + str(BalanceTracker + 6) + ',3558,0),IF($C$22>=' + str(BalanceTracker + 8) + ',3558,0))'
                    if engine_counter == 8:
                        test_cell = '=IF($C$22>=' + str(BalanceTracker + 6) + ',3558,0)'
                    ws[DAE11list[DAE11counter]] = test_cell
                    DAE11counter = DAE11counter + 1
                BalanceTracker = BalanceTracker + (DAECount / 2)
                remainingslots = remainingslots - 4
            BalanceTracker = 0
            remainingslots = 119
            DAECounter = DAECounter + 1
                #print ident.text, interface.text, tid.text, megabytes.text

        if DAECounter == 12:
            if DAE12counter != 0:
                DAEnumber = 12
            print 'there are ' + str(DAE12counter) + ' disks in DAE12'
            # fill the rest of the DAE before going to the next one
            remainingslots = remainingslots - DAE12counter
            while remainingslots > 4:
                for x in range(0, 4):
                    #print "We're on time %d" % (x)
                    if engine_counter == 6:
                        test_cell = '=IF(($C$7+$C$19=' + str(EngineCount) + '),IF($C$22>=' + str(BalanceTracker + 6) + ',3558,0),IF($C$22>=' + str(BalanceTracker + 8) + ',3558,0))'
                    if engine_counter == 8:
                        test_cell = '=IF($C$22>=' + str(BalanceTracker + 6) + ',3558,0)'
                    ws[DAE12list[DAE12counter]] = test_cell
                    DAE12counter = DAE12counter + 1
                BalanceTracker = BalanceTracker + (DAECount / 2)
                remainingslots = remainingslots - 4
            BalanceTracker = 0
            remainingslots = 119
            DAECounter = DAECounter + 1
                #print ident.text, interface.text, tid.text, megabytes.text

        if DAECounter == 13:
            if DAE13counter != 0:
                DAEnumber = 13
            print 'there are ' + str(DAE13counter) + ' disks in DAE13'
            # fill the rest of the DAE before going to the next one
            remainingslots = remainingslots - DAE13counter
            #print 'slots ' + str(remainingslots)
            while remainingslots > 4:
                while DAE13counter < 120:
                    #print 'slots ' + str(remainingslots)
                    for x in range(0, 4):
                        #print "We're on time %d" % (x)
                        if engine_counter == 6:
                            #first 8 cells need to be this
                            if DAE13counter == 0:
                                while DAE13counter <=7:
                                    test_cell = '=IF($C$19=0,0,3558)'
                                    ws[DAE13list[DAE13counter]] = test_cell
                                    DAE13counter = DAE13counter + 1
                            test_cell = '=IF($C$19=0,0,IF($C$22>=' + str(BalanceTracker + 7) + ',3558,0))'
                            ws[DAE13list[DAE13counter]] = test_cell
                        if engine_counter == 8:
                            test_cell = '=IF($C$7>=6,0,IF($C$22>=' + str(BalanceTracker + 7) + ',3558,0))'
                            ws[DAE13list[DAE13counter]] = test_cell
                        DAE13counter = DAE13counter + 1
                    BalanceTracker = BalanceTracker + (DAECount / 2)
                remainingslots = remainingslots - 4
                    #print remainingslots
            BalanceTracker = 0
            remainingslots = 119
            DAECounter = DAECounter + 1
            #print ident.text, interface.text, tid.text, megabytes.text

        if DAECounter == 14:
            if DAE14counter != 0:
                DAEnumber = 14
            print 'there are ' + str(DAE14counter) + ' disks in DAE14'
            # fill the rest of the DAE before going to the next one
            remainingslots = remainingslots - DAE14counter
            #print 'slots ' + str(remainingslots)
            while remainingslots > 4:
                while DAE14counter < 120:
                    #print 'slots ' + str(remainingslots)
                    for x in range(0, 4):
                        #print "We're on time %d" % (x)
                        if engine_counter == 6:
                            #first 8 cells need to be this
                            if DAE14counter == 0:
                                while DAE14counter <=7:
                                    test_cell = '=IF($C$19=0,0,3558)'
                                    ws[DAE14list[DAE14counter]] = test_cell
                                    DAE14counter = DAE14counter + 1
                            test_cell = '=IF($C$19=0,0,IF($C$22>=' + str(BalanceTracker + 7) + ',3558,0))'
                            ws[DAE14list[DAE14counter]] = test_cell
                        if engine_counter == 8:
                            test_cell = '=IF($C$7>=6,0,IF($C$22>=' + str(BalanceTracker + 7) + ',3558,0))'
                            ws[DAE14list[DAE14counter]] = test_cell
                        DAE14counter = DAE14counter + 1
                    BalanceTracker = BalanceTracker + (DAECount / 2)
                remainingslots = remainingslots - 4
                    #print remainingslots
            BalanceTracker = 0
            remainingslots = 119
            DAECounter = DAECounter + 1
            #print ident.text, interface.text, tid.text, megabytes.text

        if DAECounter == 15:
            if DAE15counter != 0:
                DAEnumber = 15
            print 'there are ' + str(DAE15counter) + ' disks in DAE15'
            # fill the rest of the DAE before going to the next one
            remainingslots = remainingslots - DAE15counter
            #print 'slots ' + str(remainingslots)
            while remainingslots > 4:
                while DAE15counter < 120:
                    #print 'slots ' + str(remainingslots)
                    for x in range(0, 4):
                        #print "We're on time %d" % (x)
                        if engine_counter == 6:
                            #first 8 cells need to be this
                            if DAE15counter == 0:
                                while DAE15counter <=7:
                                    test_cell = '=IF($C$19=0,0,3558)'
                                    ws[DAE15list[DAE15counter]] = test_cell
                                    DAE15counter = DAE15counter + 1
                            test_cell = '=IF($C$19=0,0,IF($C$22>=' + str(BalanceTracker + 7) + ',3558,0))'
                            ws[DAE15list[DAE15counter]] = test_cell
                        if engine_counter == 8:
                            test_cell = '=IF($C$7>=6,0,IF($C$22>=' + str(BalanceTracker + 7) + ',3558,0))'
                            ws[DAE15list[DAE15counter]] = test_cell
                        DAE15counter = DAE15counter + 1
                    BalanceTracker = BalanceTracker + (DAECount / 2)
                remainingslots = remainingslots - 4
                    #print remainingslots
            BalanceTracker = 0
            remainingslots = 119
            DAECounter = DAECounter + 1
            #print ident.text, interface.text, tid.text, megabytes.text

        if DAECounter == 16:
            if DAE16counter != 0:
                DAEnumber = 16
            print 'there are ' + str(DAE16counter) + ' disks in DAE16'
            # fill the rest of the DAE before going to the next one
            remainingslots = remainingslots - DAE16counter
            #print 'slots ' + str(remainingslots)
            while remainingslots > 4:
                while DAE16counter < 120:
                    #print 'slots ' + str(remainingslots)
                    for x in range(0, 4):
                        #print "We're on time %d" % (x)
                        if engine_counter == 6:
                            #first 8 cells need to be this
                            if DAE16counter == 0:
                                while DAE16counter <=7:
                                    test_cell = '=IF($C$19=0,0,3558)'
                                    ws[DAE16list[DAE16counter]] = test_cell
                                    DAE16counter = DAE16counter + 1
                            test_cell = '=IF($C$19=0,0,IF($C$22>=' + str(BalanceTracker + 7) + ',3558,0))'
                            ws[DAE16list[DAE16counter]] = test_cell
                        if engine_counter == 8:
                            test_cell = '=IF($C$7>=6,0,IF($C$22>=' + str(BalanceTracker + 7) + ',3558,0))'
                            ws[DAE16list[DAE16counter]] = test_cell
                        DAE16counter = DAE16counter + 1
                    BalanceTracker = BalanceTracker + (DAECount / 2)
                remainingslots = remainingslots - 4
                    #print remainingslots
            BalanceTracker = 0
            remainingslots = 119
            DAECounter = DAECounter + 1
            #print ident.text, interface.text, tid.text, megabytes.text

    # populate the spare drives

    spare_total = len(l)
    print 'spare total = ' + str(spare_total)
    DAEslot = 0
    test_cell = 'S'
    while spare_total > 0:
        if spare_total > 0:
            ws[DAE1list[DAEslot]] = test_cell
            spare_total = spare_total - 1
        if spare_total > 0:
            ws[DAE2list[DAEslot]] = test_cell
            spare_total = spare_total - 1
        if spare_total > 0:
            ws[DAE3list[DAEslot]] = test_cell
            spare_total = spare_total - 1
        if spare_total > 0:
            ws[DAE4list[DAEslot]] = test_cell
            spare_total = spare_total - 1
        if spare_total > 0:
            ws[DAE5list[DAEslot]] = test_cell
            spare_total = spare_total - 1
        if spare_total > 0:
            ws[DAE6list[DAEslot]] = test_cell
            spare_total = spare_total - 1
        if spare_total > 0:
            ws[DAE7list[DAEslot]] = test_cell
            spare_total = spare_total - 1
        if spare_total > 0:
            ws[DAE8list[DAEslot]] = test_cell
            spare_total = spare_total - 1
        if spare_total > 0:
            ws[DAE9list[DAEslot]] = test_cell
            spare_total = spare_total - 1
        if spare_total > 0:
            ws[DAE10list[DAEslot]] = test_cell
            spare_total = spare_total - 1
        if spare_total > 0:
            ws[DAE11list[DAEslot]] = test_cell
            spare_total = spare_total - 1
        if spare_total > 0:
            ws[DAE12list[DAEslot]] = test_cell
            spare_total = spare_total - 1
        if spare_total > 0:
            ws[DAE13list[DAEslot]] = test_cell
            spare_total = spare_total - 1
        if spare_total > 0:
            ws[DAE14list[DAEslot]] = test_cell
            spare_total = spare_total - 1
        if spare_total > 0:
            ws[DAE15list[DAEslot]] = test_cell
            spare_total = spare_total - 1
        if spare_total > 0:
            ws[DAE16list[DAEslot]] = test_cell
            spare_total = spare_total - 1
        DAEslot = DAEslot + 1
    
    
    ws['C8'] = DAEnumber
    #print DAEnumber
    # Save the file
    filename = 'C:\\Users\\kjensb\\calculator\\' + str(serialnumber) + ' CALCULATOR.xlsx'
    wb2.save(filename)

else:
    print 'SYMCLI failed...'

